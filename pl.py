import openpyxl
path = "/home/karthikeyan/vscode/python/sample.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
target=0
loss=0
for i in range(2,sheet_obj.max_row-3):
    if((str(sheet_obj["J"+str(i)].value))==str(1)):
        if((str(sheet_obj["K"+str(i)].value))==str(1)):
            target=target+((float(sheet_obj["C"+str(i)].value)-float(sheet_obj["D"+str(i)].value))*3)
        elif((str(sheet_obj["L"+str(i)].value))==str(1)):
            loss=loss+(float(sheet_obj["C"+str(i)].value)-float(sheet_obj["D"+str(i)].value))
sheet_obj["K"+str(sheet_obj.max_row+3)].value=target
sheet_obj["L"+str(sheet_obj.max_row+3)].value=loss
wb_obj.save(path)