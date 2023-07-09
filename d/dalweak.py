import openpyxl
import load
path = load.path
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
target=0
loss=0
n=0
countp=0
countl=0
for i in range(2,sheet_obj.max_row-3):
    n=n+1
    if((str(sheet_obj["J"+str(i)].value))==str(1)):
        if((str(sheet_obj["K"+str(i)].value))==str(1)):
            target=target+(((float(sheet_obj["C"+str(i)].value)-float(sheet_obj["D"+str(i)].value))*3))
        elif((str(sheet_obj["L"+str(i)].value))==str(1)):
            loss=loss+((float(sheet_obj["C"+str(i)].value)-float(sheet_obj["D"+str(i)].value))+1.5)
    if(n==75):
        sheet_obj["M"+str(i)].value=target
        sheet_obj["N"+str(i)].value=loss
        if(target>=loss):
           sheet_obj["O"+str(i)].value="profit" 
           countp=countp+1
        else:
            sheet_obj["O"+str(i)].value="loss"
            countl=countl+1
        n=0
        target=0
        loss=0
sheet_obj["M"+str(sheet_obj.max_row+3)].value=countp
sheet_obj["N"+str(sheet_obj.max_row+3)].value=countl      
wb_obj.save(path)