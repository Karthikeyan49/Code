import openpyxl
path = "/home/karthikeyan/vscode/python/5-15 TO 7-8.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
for i in range(2,sheet_obj.max_row-3):
    if((str(sheet_obj["J"+str(i)].value))==str(1)):
        for j in range(i+1,sheet_obj.max_row-3):
            if(float(sheet_obj["D"+str(i)].value)>float(sheet_obj["D"+str(j)].value)):
                   if(float(sheet_obj["D"+str(i)].value)>float(sheet_obj["E"+str(j)].value)):
                       if(str(sheet_obj["A"+str(i)].value)[11:19]==str("09:15:00")):
                           continue
                       else:
                            sheet_obj["L"+str(i)].value=1
                            break
            elif(float(sheet_obj["C"+str(j)].value)>=(((float(sheet_obj["C"+str(i)].value)-(float(sheet_obj["D"+str(i)].value)))*3)+(float(sheet_obj["C"+str(i)].value)))):
                sheet_obj["K"+str(i)].value=1
                break
wb_obj.save(path)